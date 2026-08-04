"""Does keyword growth predict traffic growth? — the Rising hypothesis, tested.

Runs entirely on SEMrush Organic Pages exports already on disk. NO API units.

    predictor   dOr  (export A -> export B)
    outcome     dOt  (export B -> export C)
    null        dOt  (export A -> export B) -> same outcome

`Or` must BEAT THE NULL to mean anything: traffic's own past change already predicts
its future to some degree, so the question is whether keywords add information over
it. Reported as a partial correlation controlling for the null.

The trap this design avoids: taking known breakouts and confirming that keywords rose
first. That conditions on the outcome and porridge will always look convincing. Here
the cohort is EVERY page present in all three exports, so the pages where keywords rose
and nothing happened are counted too.

First run (2026-08-04, seriouseats.com, Jun 21 / Jul 21 / Aug 03, n=5,064):

    corr(dOr, outcome)              +0.1162
    corr(dOt, outcome)              -0.2330   <- traffic momentum is INVERTED
    partial corr(dOr | dOt)         +0.1865
    decile 1 -> 10 by dOr:          29% grew -> 51% grew, monotonic
    dOr > +0.10:  52% grew vs 43% baseline

So the hypothesis survives in a WEAK form — Or is a legitimate ranking signal for a
watchlist, explaining ~3-4% of variance, not a confident breakout detector. The more
valuable finding is the null: ranking risers by traffic change would do WORSE than
random.

Usage:
    python -m scripts.keyword_lead_experiment <A.xlsx> <B.xlsx> <C.xlsx>
    python -m scripts.keyword_lead_experiment            # defaults to the seriouseats trio
"""
from __future__ import annotations

import math
import statistics as st
import sys
from pathlib import Path

from openpyxl import load_workbook

DEFAULTS = [
    "input/seriouseats.com-organic.PagesV3-us-20260621-2026-06-22T22_29_17Z.xlsx",
    "input/seriouseats.com-organic.PagesV3-us-20260721-2026-07-22T15_21_27Z.xlsx",
    r"C:\Users\john\Downloads\seriouseats.com-organic.PagesV3-us-20260803-2026-08-04T16_56_24Z.xlsx",
]


def load(path: str) -> dict:
    """{url: (traffic, traffic_change, keyword_count)} from an Organic Pages export."""
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    h = rows[0]
    iU, iT, iC, iK = (h.index("URL"), h.index("Traffic"),
                      h.index("Traffic Change"), h.index("Number of Keywords"))
    out = {r[iU]: (r[iT] or 0, r[iC] or 0, r[iK] or 0) for r in rows[1:] if r[iU]}
    wb.close()
    return out


def corr(xs, ys) -> float:
    n = len(xs)
    if n < 3:
        return 0.0
    mx, my = sum(xs) / n, sum(ys) / n
    sxy = sum((a - mx) * (b - my) for a, b in zip(xs, ys))
    den = (sum((a - mx) ** 2 for a in xs) * sum((b - my) ** 2 for b in ys)) ** 0.5
    return sxy / den if den else 0.0


def main() -> int:
    paths = sys.argv[1:] or DEFAULTS
    if len(paths) != 3:
        print("need exactly three exports, oldest first"); return 1
    for p in paths:
        if not Path(p).exists():
            print(f"missing: {p}"); return 1

    A, B, C = (load(p) for p in paths)
    print(f"rows   A {len(A):,}   B {len(B):,}   C {len(C):,}")
    urls = set(A) & set(B) & set(C)
    print(f"present in ALL THREE: {len(urls):,}\n")
    if len(urls) < 100:
        print("too few overlapping urls to be worth reading"); return 1

    lg = lambda v: math.log10(v + 1)          # noqa: E731 - log so big and small pages compare
    X_or, X_ot, Y = [], [], []
    for u in urls:
        ot_a, _, or_a = A[u]
        ot_b, _, or_b = B[u]
        ot_c, _, _ = C[u]
        X_or.append(lg(or_b) - lg(or_a))      # predictor
        X_ot.append(lg(ot_b) - lg(ot_a))      # null
        Y.append(lg(ot_c) - lg(ot_b))         # outcome

    c_or, c_ot, r23 = corr(X_or, Y), corr(X_ot, Y), corr(X_or, X_ot)
    den = ((1 - c_ot ** 2) * (1 - r23 ** 2)) ** 0.5
    partial = (c_or - c_ot * r23) / den if den else 0.0

    print("PREDICTIVE TEST")
    print(f"   corr( dOr , outcome )            = {c_or:+.4f}   the hypothesis")
    print(f"   corr( dOt , outcome )            = {c_ot:+.4f}   the null")
    print(f"   difference                        = {c_or - c_ot:+.4f}")
    print(f"   PARTIAL corr( dOr | dOt )        = {partial:+.4f}   unique contribution")

    print("\nDOES IT WORK AT THE TOP? deciles by dOr:")
    pairs = sorted(zip(X_or, Y), key=lambda p: p[0])
    n, k = len(pairs), 10
    print(f"   {'decile':<10}{'n':>6}{'mean outcome':>15}{'% grew':>10}")
    for i in range(k):
        chunk = pairs[i * n // k:(i + 1) * n // k]
        ys = [y for _, y in chunk]
        print(f"   {i+1:<10}{len(chunk):>6}{st.mean(ys):>+15.4f}"
              f"{sum(1 for y in ys if y > 0) / len(ys) * 100:>9.0f}%")

    top = [y for x, y in pairs if x > 0.10]
    if top:
        print(f"\n   dOr > +0.10 (~+26%/mo): n={len(top)}, mean {st.mean(top):+.4f} "
              f"vs corpus {st.mean(Y):+.4f}")
        print(f"   {sum(1 for y in top if y > 0)/len(top)*100:.0f}% grew "
              f"vs {sum(1 for y in Y if y > 0)/len(Y)*100:.0f}% baseline")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
